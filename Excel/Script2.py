from itertools import islice
from openpyxl import load_workbook, Workbook

#Load in the workbook

wb = load_workbook('test.xlsx')

#Get sheet names
print(wb.get_sheet_names())



#Get a sheet by name
sheet = wb.get_sheet_by_name('Sheet3')

#Print the sheet title
print(sheet.title)

#Get currently active sheet
anotherSheet = wb.active

#Check `anotherSheet`
print(anotherSheet)


#########################################################


#Retrieve the value of a certain cell
print(sheet['A1'].value)

#Select element 'B2' of your sheet
c = sheet['B2']

#Retrieve the row number of your element
print('row : ' + str(c.row))

#Retrieve the column letter of your element
print('column : ' + str(c.column))

#Retrieve the coordinates of the cell
print("coord : " + str(c.coordinate))




#Retrieve cell value
print(sheet.cell(row = 1, column = 2).value)


#Print out values in column 2

for i in range(1, 4):
    print(i, sheet.cell(row = i, column = 2).value)


#################################################

from openpyxl.utils import get_column_letter, column_index_from_string

#Return 'A'
print(get_column_letter(1))

#Return '1'
print(column_index_from_string('A'))



#Print row per row
for cellObj in sheet['A1':'C3']:
    for cell in cellObj:
        print(cell.coordinate, cell.value)

    print('----END----')



#Retrieve the maximum amount of rows
print(sheet.max_row)

#Retrieve the maximum amount of column
print(sheet.max_column)



############################################

import pandas as pd

#Convert Sheet to DataFrame
df = pd.DataFrame(sheet.values)

#Put the sheet values in `data`
data = sheet.values

#Indicate the columns in the sheet values
cols = next(data)[1:]

#Convert your data to a list
data = list(data)

idx = [r[0] for r in data]

#Slice the data at index 1

data = (islice(r, 1, None)for r in data)

#Make your DataFrame
df = pd.DataFrame(data, index = idx, columns = cols)



#Import `dataframe_to_rows`
from openpyxl.utils.dataframe import  dataframe_to_rows

#Initialize a workbook
wb = Workbook()

#Get the workshet in the active workbook
ws = wb.active


#Append the rows of the DataFrame to your worksheet
for r in dataframe_to_rows(df, index = True, header = True):
    ws.append(r)




#Для форматирования данных в Excel
import xlrd


#Open a workbook
workbook = xlrd.open_workbook('test.xls')

#Load only current sheets to memory

workbook = xlrd.open_workbook('test.xls', on_demand = True)


#Load a specific sheet by name
worksheet = workbook.sheet_by_name('Sheet1')

#Load a specific sheet by index
worksheet = workbook.sheet_by_index(0)

#Retrieve the value from cell at indices (0, 0)
print(sheet.cell(0, 0).value)


########################################################

import xlwt

#Initialize a workbook
book = xlwt.Workbook(encoding = 'utf-8')

#Add a sheet to the workbook
sheet1 = book.add_sheet('Python Sheet 1')

#Write to the sheet of the workbook
sheet1.write(0, 0, "This is the First Cell of the First Sheet")

#Save the workbook
book.save("spreadsheet.xls")


#Initialize a workbook
book = xlwt.Workbook()

#Add a sheet to the workbook
sheet1 = book.add_sheet("Sheet1")


#The data
cols = ["A", "B", "C", "D", "E"]
txt = [0, 1, 2, 3, 4]

#Loop over the rows and columns and fill in the values
for num in range(5):
    row = sheet1.row(num)
    for index, col in enumerate(cols):
        value = txt[index] + num
        row.write(index, value)


#Save the result
book.save("test.xls")


#pyexcel

import pyexcel

#Get an array from the data
my_array = pyexcel.get_array(file_name = "test.xls")

#Для получения данных в упорядоченном словаре списков
from pyexcel._compact import OrderedDict

#Get your data in an ordered dictionary of lists
my_dict = pyexcel.get_dict(file_name = "test.xls", name_columns_by_row = 0)

#Get your data in a dictionary of 2D arrays
book_dict = pyexcel.get_book_dict(file_name = "test.xls")




#Retrieve the records of the file
records = pyexcel.get_records(file_name = "test.xls")



#Запись файлов
#Get the data
data =[[1, 2, 3], [4, 5, 6], [7, 8, 9]]

#Save the array to a file
pyexcel.save_as(array = data, dest_file_name = 'array_data.xls')



#The data

a2d_array_dictionary = {'Sheet 1': [
                                   ['ID', 'AGE', 'SCORE']
                                   [1, 22, 5],
                                   [2, 15, 6],
                                   [3, 28, 9]
                                  ],
                       'Sheet 2': [
                                    ['X', 'Y', 'Z'],
                                    [1, 2, 3],
                                    [4, 5, 6]
                                    [7, 8, 9]
                                  ],
                       'Sheet 3': [
                                    ['M', 'N', 'O', 'P'],
                                    [10, 11, 12, 13],
                                    [14, 15, 16, 17]
                                    [18, 19, 20, 21]
                                   ]}

#Save the data to a file
pyexcel.save_book_as(bookdict = a2d_array_dictionary, dest_file_name = 'a2d_array_data.xls')

